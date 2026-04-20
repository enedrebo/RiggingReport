from datetime import datetime
import openpyxl


# ── Helpers ──────────────────────────────────────────────────────────────────

def _txt(ws, addr):
    """Return cell value as a stripped string (empty string if None)."""
    v = ws[addr].value
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d.%m.%Y")
    return str(v).strip()


def _num(ws, addr):
    """Return cell value as float, or None if the cell is empty."""
    v = ws[addr].value
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _require_num(ws, addr, label):
    """Like _num but raises ValueError if the cell is empty or negative."""
    v = _num(ws, addr)
    if v is None:
        raise ValueError(f"Required value missing in cell {addr} ({label})")
    if v < 0:
        raise ValueError(f"Value in cell {addr} ({label}) must be >= 0, got {v}")
    return v


def fmt(v):
    """Format a number for display (5 significant figures, no trailing zeros)."""
    if v is None:
        return ""
    return f"{v:.5g}"


_LOAD_HTML = {
    "F_hook": "F<sub>hook</sub>",
    "F_leg":  "F<sub>leg</sub>",
    "F_leg1": "F<sub>leg1</sub>",
    "F_leg2": "F<sub>leg2</sub>",
    "F_leg3": "F<sub>leg3</sub>",
    "F_leg4": "F<sub>leg4</sub>",
}


def _load_html(label):
    return _LOAD_HTML.get(label, label)


# ── Sheet readers ─────────────────────────────────────────────────────────────

def _read_loads(ws):
    date_val = ws["C5"].value
    date_str = (date_val.strftime("%d.%m.%Y")
                if isinstance(date_val, datetime) else str(date_val or ""))

    # Static hook load
    w_lift    = _require_num(ws, "D13", "Weight of object")
    y_con     = _require_num(ws, "D14", "Weight inaccuracy factor")
    w_rig     = _require_num(ws, "D15", "Weight of lift rigging")
    w_special = _require_num(ws, "D16", "Special weights")
    shl       = _require_num(ws, "D17", "Static hook load")

    static_rows = [
        {"desc": _txt(ws, "B13"),
         "formula": f"W<sub>lift</sub> = {fmt(w_lift)}",
         "comment": _txt(ws, "E13")},
        {"desc": _txt(ws, "B14"),
         "formula": f"γ<sub>con</sub> = {fmt(y_con)}",
         "comment": _txt(ws, "E14")},
        {"desc": _txt(ws, "B15"),
         "formula": f"W<sub>rig</sub> = {fmt(w_rig)}",
         "comment": _txt(ws, "E15")},
        {"desc": _txt(ws, "B16"),
         "formula": f"W<sub>special</sub> = {fmt(w_special)}",
         "comment": _txt(ws, "E16")},
        {"desc": _txt(ws, "B17"),
         "formula": (f"SHL = W<sub>lift</sub> × γ<sub>con</sub>"
                     f" + W<sub>rig</sub> + W<sub>special</sub> = {fmt(shl)}"),
         "comment": _txt(ws, "E17")},
    ]

    # Dynamic hook load — prefer D20 (manual input), fall back to D21 (table)
    daf = _num(ws, "D20")
    daf_comment = _txt(ws, "E20")
    daf_desc    = _txt(ws, "B20")
    if daf is None:
        daf         = _num(ws, "D21")
        daf_comment = _txt(ws, "E21")
        daf_desc    = _txt(ws, "B21")
    if daf is None:
        raise ValueError("DAF not found in D20 or D21")

    f_hook = _require_num(ws, "D22", "Dynamic hook load")

    dynamic_rows = [
        {"desc": daf_desc,
         "formula": f"DAF = {fmt(daf)}",
         "comment": daf_comment},
        {"desc": _txt(ws, "B22"),
         "formula": f"F<sub>hook</sub> = SHL × DAF = {fmt(f_hook)}",
         "comment": _txt(ws, "E22")},
    ]

    # Sling loads: direct analysis results take priority over calculated
    legs = [
        ("F<sub>leg1</sub>", _num(ws, "D26"), _txt(ws, "E26"), _txt(ws, "B26")),
        ("F<sub>leg2</sub>", _num(ws, "D27"), _txt(ws, "E27"), _txt(ws, "B27")),
        ("F<sub>leg3</sub>", _num(ws, "D28"), _txt(ws, "E28"), _txt(ws, "B28")),
        ("F<sub>leg4</sub>", _num(ws, "D29"), _txt(ws, "E29"), _txt(ws, "B29")),
    ]

    if any(val is not None for _, val, _, _ in legs):
        sling_title = "Sling loads from analysis"
        sling_rows = [
            {"desc": desc, "formula": f"{lbl} = {fmt(val)}", "comment": cmt}
            for lbl, val, cmt, desc in legs if val is not None
        ]
    else:
        n_leg   = _require_num(ws, "D31", "Leg quantity")
        leg_ang = _require_num(ws, "D32", "Leg angle")
        cog_sf  = _require_num(ws, "D33", "CoG shift factor")
        skl     = _require_num(ws, "D34", "Skew load factor")
        f_leg   = _require_num(ws, "D35", "Leg max. dynamic load")
        sling_title = "Sling loads"
        sling_rows = [
            {"desc": _txt(ws, "B31"),
             "formula": f"n<sub>leg</sub> = {fmt(n_leg)}",
             "comment": _txt(ws, "E31")},
            {"desc": _txt(ws, "B32"),
             "formula": f"β<sub>leg</sub> = {fmt(leg_ang)}°",
             "comment": _txt(ws, "E32")},
            {"desc": _txt(ws, "B33"),
             "formula": f"CoG<sub>sf</sub> = {fmt(cog_sf)}",
             "comment": _txt(ws, "E33")},
            {"desc": _txt(ws, "B34"),
             "formula": f"SKL = {fmt(skl)}",
             "comment": _txt(ws, "E34")},
            {"desc": _txt(ws, "B35"),
             "formula": (f"F<sub>leg</sub> = (F<sub>hook</sub> × CoG<sub>sf</sub>"
                         f" × SKL) / (n<sub>leg</sub> × sin(β<sub>leg</sub>))"
                         f" = {fmt(f_leg)}"),
             "comment": _txt(ws, "E35")},
        ]

    return {
        "project":  _txt(ws, "C3"),
        "date":     date_str,
        "author":   _txt(ws, "C6"),
        "approver": _txt(ws, "C7"),
        "title":    _txt(ws, "C4"),
        "image":    _txt(ws, "C10"),
        "comment":  _txt(ws, "B38"),
        "sections": [
            {"title": "Static hook load",  "rows": static_rows},
            {"title": "Dynamic hook load", "rows": dynamic_rows},
            {"title": sling_title,         "rows": sling_rows},
        ],
    }


def _read_slings(ws):
    item_no  = _txt(ws, "D4")
    item_desc = _txt(ws, "D5")
    det_desc  = _txt(ws, "D6")

    mbl      = _require_num(ws, "D12", "MBL")
    material = _txt(ws, "D13")
    terminat = _txt(ws, "D14")
    d_bend   = _num(ws, "D15")
    d_item   = _num(ws, "D16")
    d_over_d = _num(ws, "D17")

    general_rows = [
        {"desc": _txt(ws, "B12"), "formula": f"MBL = {fmt(mbl)}",        "comment": _txt(ws, "E12")},
        {"desc": _txt(ws, "B13"), "formula": material,                    "comment": _txt(ws, "E13")},
        {"desc": _txt(ws, "B14"), "formula": terminat,                    "comment": _txt(ws, "E14")},
    ]
    if d_bend is not None:
        general_rows.append({"desc": _txt(ws, "B15"), "formula": f"D = {fmt(d_bend)}",     "comment": _txt(ws, "E15")})
    if d_item is not None:
        general_rows.append({"desc": _txt(ws, "B16"), "formula": f"d = {fmt(d_item)}",     "comment": _txt(ws, "E16")})
    if d_over_d is not None:
        general_rows.append({"desc": _txt(ws, "B17"), "formula": f"D/d = {fmt(d_over_d)}", "comment": _txt(ws, "E17")})

    yb   = _require_num(ws, "D20", "γb")
    ys   = _require_num(ws, "D21", "γs")
    yr   = _require_num(ws, "D22", "γr")
    yh   = _require_num(ws, "D23", "γh")
    yc   = _require_num(ws, "D24", "γc")
    ym   = _require_num(ws, "D25", "γm")
    yw   = _require_num(ws, "D26", "γw")
    ysf1 = _require_num(ws, "D27", "γsf1")
    ysf2 = _require_num(ws, "D28", "γsf2")
    sf   = _require_num(ws, "D29", "SF")

    sf_rows = [
        {"desc": _txt(ws, "B20"),
         "formula": f"γ<sub>b</sub> = {fmt(yb)}",
         "comment": _txt(ws, "E20")},
        {"desc": _txt(ws, "B21"),
         "formula": f"γ<sub>s</sub> = {fmt(ys)}",
         "comment": _txt(ws, "E21")},
        {"desc": _txt(ws, "B22"),
         "formula": f"γ<sub>r</sub> = Max(γ<sub>b</sub>, γ<sub>req</sub>) = {fmt(yr)}",
         "comment": _txt(ws, "E22")},
        {"desc": _txt(ws, "B23"),
         "formula": f"γ<sub>h</sub> = {fmt(yh)}",
         "comment": _txt(ws, "E23")},
        {"desc": _txt(ws, "B24"),
         "formula": f"γ<sub>c</sub> = {fmt(yc)}",
         "comment": _txt(ws, "E24")},
        {"desc": _txt(ws, "B25"),
         "formula": f"γ<sub>m</sub> = {fmt(ym)}",
         "comment": _txt(ws, "E25")},
        {"desc": _txt(ws, "B26"),
         "formula": f"γ<sub>w</sub> = {fmt(yw)}",
         "comment": _txt(ws, "E26")},
        {"desc": _txt(ws, "B27"),
         "formula": (f"γ<sub>sf1</sub> = γ<sub>r</sub> × γ<sub>h</sub>"
                     f" × γ<sub>c</sub> × γ<sub>m</sub> × γ<sub>w</sub> = {fmt(ysf1)}"),
         "comment": _txt(ws, "E27")},
        {"desc": _txt(ws, "B28"),
         "formula": f"γ<sub>sf2</sub> = 2.3 × γ<sub>r</sub> × γ<sub>w</sub> = {fmt(ysf2)}",
         "comment": _txt(ws, "E28")},
        {"desc": _txt(ws, "B29"),
         "formula": f"γ<sub>sf</sub> = Max(γ<sub>sf1</sub>, γ<sub>sf2</sub>) = {fmt(sf)}",
         "comment": _txt(ws, "E29")},
    ]

    max_dyn_type = _txt(ws, "C32") or "F_leg"
    mdl     = _load_html(max_dyn_type)
    max_dyn = _require_num(ws, "D32", "Max. dynamic load")
    ylf     = _require_num(ws, "D33", "Shape factor")
    mbl_r   = _require_num(ws, "D34", "Required MBL")

    load_rows = [
        {"desc": _txt(ws, "B32"),
         "formula": f"{mdl} = {fmt(max_dyn)}",
         "comment": _txt(ws, "E32")},
        {"desc": _txt(ws, "B33"),
         "formula": f"γ<sub>LF</sub> = {fmt(ylf)}",
         "comment": _txt(ws, "E33")},
        {"desc": _txt(ws, "B34"),
         "formula": (f"MBL<sub>req</sub> = (MBL × γ<sub>LF</sub>)"
                     f" / γ<sub>sf</sub> = {fmt(mbl_r)}"),
         "comment": _txt(ws, "E34")},
    ]

    uf = _require_num(ws, "D36", "Utilization ratio")
    uf_rows = [
        {"desc": _txt(ws, "B36"),
         "formula": f"UF = {mdl} / MBL<sub>req</sub> = {fmt(uf)}",
         "comment": _txt(ws, "E36")},
    ]

    return {
        "title_line1": f"Item no. {item_no} - {item_desc}",
        "title_line2": det_desc,
        "item_no":   item_no,
        "item_desc": item_desc,
        "det_desc":  det_desc,
        "uf":        uf,
        "image":   _txt(ws, "D9"),
        "comment": _txt(ws, "B39"),
        "sections": [
            {"title": "General",        "rows": general_rows},
            {"title": "Safety factors", "rows": sf_rows},
            {"title": "Loads",          "rows": load_rows},
            {"title": "Utilization",    "rows": uf_rows},
        ],
    }


def _read_shackles(ws):
    item_no   = _txt(ws, "D4")
    item_desc = _txt(ws, "D5")
    det_desc  = _txt(ws, "D6")

    wll     = _require_num(ws, "D12", "WLL")
    item_sf = _require_num(ws, "D13", "Item SF")
    mbl     = _require_num(ws, "D14", "MBL")

    general_rows = [
        {"desc": _txt(ws, "B12"),
         "formula": f"WLL = {fmt(wll)}",
         "comment": _txt(ws, "E12")},
        {"desc": _txt(ws, "B13"),
         "formula": f"SF<sub>item</sub> = {fmt(item_sf)}",
         "comment": _txt(ws, "E13")},
        {"desc": _txt(ws, "B14"),
         "formula": f"MBL = {fmt(mbl)}",
         "comment": _txt(ws, "E14")},
    ]

    wll_daf   = _require_num(ws, "D17", "WLL × DAF")
    mbl_3     = _require_num(ws, "D18", "MBL/3.0")
    dpl       = _require_num(ws, "D19", "Documented proof load")
    allowable = _require_num(ws, "D20", "Allowable load")

    sf_rows = [
        {"desc": _txt(ws, "B17"),
         "formula": f"R1 = WLL × DAF = {fmt(wll_daf)}",
         "comment": _txt(ws, "E17")},
        {"desc": _txt(ws, "B18"),
         "formula": f"R2 = MBL / 3.0 = {fmt(mbl_3)}",
         "comment": _txt(ws, "E18")},
        {"desc": _txt(ws, "B19"),
         "formula": f"R3 = Documented proof load = {fmt(dpl)}",
         "comment": _txt(ws, "E19")},
        {"desc": _txt(ws, "B20"),
         "formula": f"Allowable load = Min(R1, R2, R3) = {fmt(allowable)}",
         "comment": _txt(ws, "E20")},
    ]

    max_dyn_type = _txt(ws, "C23") or "F_leg"
    mdl     = _load_html(max_dyn_type)
    max_dyn = _require_num(ws, "D23", "Max. dynamic load")

    load_rows = [
        {"desc": _txt(ws, "B23"),
         "formula": f"{mdl} = {fmt(max_dyn)}",
         "comment": _txt(ws, "E23")},
    ]

    uf = _require_num(ws, "D25", "Utilization ratio")
    uf_rows = [
        {"desc": _txt(ws, "B25"),
         "formula": f"UF = {mdl} / Allowable load = {fmt(uf)}",
         "comment": _txt(ws, "E25")},
    ]

    return {
        "title_line1": f"Item no. {item_no} - {item_desc}",
        "title_line2": det_desc,
        "item_no":   item_no,
        "item_desc": item_desc,
        "det_desc":  det_desc,
        "uf":        uf,
        "image":   _txt(ws, "D9"),
        "comment": _txt(ws, "B28"),
        "sections": [
            {"title": "General",        "rows": general_rows},
            {"title": "Safety factors", "rows": sf_rows},
            {"title": "Loads",          "rows": load_rows},
            {"title": "Utilization",    "rows": uf_rows},
        ],
    }


# ── Public API ────────────────────────────────────────────────────────────────

def read_page_list(source):
    """
    Lightweight read — returns list of {title, image_filename} for each report
    page, in order. Used by the Streamlit UI to build the image management table.
    source can be a file path string or a BytesIO object.
    """
    wb = openpyxl.load_workbook(source, data_only=True)
    pages = []

    if "0-Loads" in wb.sheetnames:
        ws = wb["0-Loads"]
        pages.append({
            "title": _txt(ws, "C4") or "Loads",
            "image_filename": _txt(ws, "C10"),
        })

    for name in wb.sheetnames:
        if name in ("0-Loads", "Instructions"):
            continue
        ws = wb[name]
        sheet_type = ws["A1"].value
        if sheet_type == _SLINGS_TYPE:
            item_no   = _txt(ws, "D4")
            item_desc = _txt(ws, "D5")
            det_desc  = _txt(ws, "D6")
            uf        = _num(ws, "D36")
            pages.append({
                "title": f"Item no. {item_no} - {item_desc}",
                "image_filename": _txt(ws, "D9"),
                "item_no": item_no,
                "item_desc": item_desc,
                "det_desc": det_desc,
                "uf": uf,
            })
        elif sheet_type == _SHACKLES_TYPE:
            item_no   = _txt(ws, "D4")
            item_desc = _txt(ws, "D5")
            det_desc  = _txt(ws, "D6")
            uf        = _num(ws, "D25")
            pages.append({
                "title": f"Item no. {item_no} - {item_desc}",
                "image_filename": _txt(ws, "D9"),
                "item_no": item_no,
                "item_desc": item_desc,
                "det_desc": det_desc,
                "uf": uf,
            })

    return pages

_SLINGS_TYPE   = "Slings, ropes, wire, chain"
_SHACKLES_TYPE = "Shackles, hooks, masterlinks"


def read_excel(filepath):
    """
    Parse a rigging calculations workbook.
    Returns (loads_data, [item_data, ...]) where each dict matches
    the format expected by report_template.build_pdf().
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if "0-Loads" not in wb.sheetnames:
        raise ValueError("Workbook is missing the required '0-Loads' sheet")

    loads_data = _read_loads(wb["0-Loads"])

    items = []
    for name in wb.sheetnames:
        if name in ("0-Loads", "Instructions"):
            continue
        ws = wb[name]
        sheet_type = ws["A1"].value
        if sheet_type == _SLINGS_TYPE:
            items.append(_read_slings(ws))
        elif sheet_type == _SHACKLES_TYPE:
            items.append(_read_shackles(ws))

    return loads_data, items
